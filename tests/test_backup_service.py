import io
import os
import tempfile

import msoffcrypto
import openpyxl
import pytest

from controllers.customer_controller import CustomerController
from database.models import CustomerStatus, Gender
from services.backup_service import run_backup


@pytest.fixture
def tmp_folder():
    with tempfile.TemporaryDirectory() as d:
        yield d


@pytest.fixture
def ctrl():
    return CustomerController()


def _open_encrypted(path: str, password: str) -> openpyxl.Workbook:
    """Decrypt the Excel file and return an openpyxl Workbook."""
    with open(path, "rb") as f:
        office_file = msoffcrypto.OfficeFile(f)
        office_file.load_key(password=password)
        buf = io.BytesIO()
        office_file.decrypt(buf)
    buf.seek(0)
    return openpyxl.load_workbook(buf)


def test_run_backup_creates_file(tmp_folder):
    path = run_backup(tmp_folder, "secret123")
    assert os.path.isfile(path)
    assert path.endswith(".xlsx")


def test_run_backup_file_is_password_protected(tmp_folder):
    path = run_backup(tmp_folder, "secret123")
    # Opening without password should raise
    with pytest.raises(Exception):
        openpyxl.load_workbook(path)


def test_run_backup_file_opens_with_correct_password(tmp_folder):
    path = run_backup(tmp_folder, "correct_pass")
    wb = _open_encrypted(path, "correct_pass")
    assert wb is not None


def test_run_backup_wrong_password_fails(tmp_folder):
    path = run_backup(tmp_folder, "correct_pass")
    with pytest.raises(Exception):
        _open_encrypted(path, "wrong_pass")


def test_run_backup_contains_customer_data(tmp_folder, ctrl):
    ctrl.create(
        "יוסי", "כהן", Gender.MALE,
        "050-1234567", "", "", "yossi@test.com",
        CustomerStatus.CUSTOMER, "test notes",
    )
    path = run_backup(tmp_folder, "pw123")
    wb = _open_encrypted(path, "pw123")
    ws = wb.active
    # Row 1 = header, row 2 = first customer
    values = [ws.cell(row=2, column=i).value for i in range(1, 4)]
    assert "יוסי" in values
    assert "כהן" in values


def test_run_backup_empty_db_produces_header_only(tmp_folder):
    path = run_backup(tmp_folder, "pw123")
    wb = _open_encrypted(path, "pw123")
    ws = wb.active
    assert ws.max_row == 1  # header only


def test_run_backup_invalid_folder_raises():
    with pytest.raises(ValueError, match="תיקיית הגיבוי"):
        run_backup("/nonexistent/path/xyz", "pw123")


def test_run_backup_empty_password_raises(tmp_folder):
    with pytest.raises(ValueError, match="סיסמת גיבוי"):
        run_backup(tmp_folder, "")
