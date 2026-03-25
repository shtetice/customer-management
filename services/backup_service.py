import io
import os
from datetime import datetime

import msoffcrypto
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from database.db import get_session
from database.models import Customer
from ui.styles import STATUS_LABELS


def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def run_backup(backup_folder: str, password: str) -> str:
    """Export all customers to a password-protected Excel file. Returns the saved path."""
    if not backup_folder or not os.path.isdir(backup_folder):
        raise ValueError(f"תיקיית הגיבוי אינה קיימת: {backup_folder}")
    if not password:
        raise ValueError("סיסמת גיבוי לא הוגדרה בהגדרות")

    # Fetch all customers
    session = get_session()
    try:
        customers = session.query(Customer).order_by(Customer.surname, Customer.name).all()
        rows = []
        for c in customers:
            phones = " | ".join(p for p in [c.phone, c.phone2, c.phone3] if p)
            dob = c.date_of_birth.strftime("%d/%m/%Y") if c.date_of_birth else ""
            status = STATUS_LABELS.get(c.status.value, c.status.value)
            rows.append([
                c.name, c.surname, phones, c.email or "",
                c.city or "", c.address or "", status, dob,
                c.notes or "",
            ])
    finally:
        session.close()

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "לקוחות"
    ws.sheet_view.rightToLeft = True

    headers = ["שם", "שם משפחה", "טלפון", "אימייל", "עיר", "כתובת", "סטטוס", "תאריך לידה", "הערות"]
    header_fill = PatternFill("solid", fgColor="2C3444")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    alt_fill = PatternFill("solid", fgColor="F0F2F5")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = _thin_border()

    ws.row_dimensions[1].height = 22

    for row_idx, row_data in enumerate(rows, 2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=(col_idx == 9))
            cell.border = _thin_border()
            if fill:
                cell.fill = fill

    # Column widths
    col_widths = [14, 14, 22, 26, 14, 22, 10, 14, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    # Save unencrypted to memory buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Encrypt with password
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"backup_customers_{timestamp}.xlsx"
    dest_path = os.path.join(backup_folder, filename)

    office_file = msoffcrypto.OfficeFile(buf)
    with open(dest_path, "wb") as out:
        office_file.encrypt(password, out)

    return dest_path
